from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
import cloudinary
import cloudinary.uploader
import qrcode
import os
import io
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'ton_secret_key_ici'

# Configuration Login
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = generate_password_hash('More@Admin2026#')

# Configuration Cloudinary
cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key = os.environ.get('CLOUDINARY_API_KEY'),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET')
)

# Configuration uploads
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    host = os.environ.get('MYSQL_HOST')
    user = os.environ.get('MYSQL_USER')
    password = os.environ.get('MYSQL_PASSWORD')
    database = os.environ.get('MYSQL_DATABASE')
    port = int(os.environ.get('MYSQL_PORT', 3306))
    return mysql.connector.connect(
        host=host,
        user=user,
        password=password,
        database=database,
        port=port
    )

# LISTE DES CATEGORIES
@app.route('/categories')
@login_required
def categories():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT c.*, COUNT(a.id) as nb_articles FROM categories c LEFT JOIN articles a ON a.categorie_id = c.id GROUP BY c.id")
    categories = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('categories.html', categories=categories)

# AJOUTER UNE CATEGORIE
@app.route('/categories/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_categorie():
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO categories (nom, description) VALUES (%s, %s)", (nom, description))
        conn.commit()
        cur.close()
        conn.close()
        flash('Catégorie ajoutée !', 'success')
        return redirect(url_for('categories'))
    return render_template('ajouter_categorie.html')

# SUPPRIMER UNE CATEGORIE
@app.route('/categories/supprimer/<int:id>')
@login_required
def supprimer_categorie(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE articles SET categorie_id = NULL WHERE categorie_id = %s", (id,))
    cur.execute("DELETE FROM categories WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Catégorie supprimée !', 'danger')
    return redirect(url_for('categories'))

# LOGIN
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD, password):
            login_user(User(1))
            return redirect(url_for('index'))
        flash('Identifiants incorrects !', 'danger')
    return render_template('login.html')

# LOGOUT
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# PAGE PRINCIPALE - Liste des articles
@app.route('/')
@login_required
def index():
    conn = get_db()
    cur = conn.cursor()
    
    search = request.args.get('search', '')
    categorie_filter = request.args.get('categorie', '')
    stock_filter = request.args.get('stock', '')

    query = """
        SELECT a.*, c.nom as categorie_nom 
        FROM articles a 
        LEFT JOIN categories c ON a.categorie_id = c.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND (a.nom LIKE %s OR a.description LIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])

    if categorie_filter:
        query += " AND a.categorie_id = %s"
        params.append(categorie_filter)

    if stock_filter == 'bas':
        query += " AND a.stock <= 5"
    elif stock_filter == 'ok':
        query += " AND a.stock > 5"

    cur.execute(query, params)
    articles = cur.fetchall()
    cur.execute("SELECT * FROM categories")
    categories = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('index.html', articles=articles, categories=categories, 
                         search=search, categorie_filter=categorie_filter, stock_filter=stock_filter)

@app.route('/imprimer_qr/<int:id>')
@login_required
def imprimer_qr(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM articles WHERE id = %s", (id,))
    article = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('imprimer_qr.html', article=article)

# AJOUTER UN ARTICLE
@app.route('/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter():
    conn = get_db()
    cur = conn.cursor()
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        prix = request.form['prix']
        stock = request.form['stock']
        categorie_id = request.form.get('categorie_id') or None

        photo_filename = None
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file(photo.filename):
                upload_result = cloudinary.uploader.upload(photo)
                photo_filename = upload_result['secure_url']

        cur.execute("""
            INSERT INTO articles (nom, description, prix, stock, photo, categorie_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (nom, description, prix, stock, photo_filename, categorie_id))
        conn.commit()
        article_id = cur.lastrowid

        base_url = os.environ.get('BASE_URL', 'http://localhost:5000')
        url_article = f"{base_url}/article/{article_id}"
        qr = qrcode.make(url_article)
        buffer = io.BytesIO()
        qr.save(buffer, format='PNG')
        buffer.seek(0)
        qr_upload = cloudinary.uploader.upload(buffer, resource_type='image', public_id=f"qr_{article_id}")
        qr_url = qr_upload['secure_url']

        cur.execute("UPDATE articles SET qr_code = %s WHERE id = %s", (qr_url, article_id))
        conn.commit()
        cur.close()
        conn.close()

        flash('Article ajouté avec succès !', 'success')
        return redirect(url_for('index'))

    cur.execute("SELECT * FROM categories")
    categories = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('ajouter.html', categories=categories)

# MODIFIER UN ARTICLE
@app.route('/modifier/<int:id>', methods=['GET', 'POST'])
@login_required
def modifier(id):
    conn = get_db()
    cur = conn.cursor()
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        prix = request.form['prix']
        stock = request.form['stock']
        categorie_id = request.form.get('categorie_id') or None

        # Récupérer les anciennes valeurs
        cur.execute("SELECT nom, description, prix, stock FROM articles WHERE id = %s", (id,))
        ancien = cur.fetchone()

        # Enregistrer les modifications
        champs = [('nom', ancien[0], nom), ('description', ancien[1], description),
                  ('prix', str(ancien[2]), prix), ('stock', str(ancien[3]), stock)]
        for champ, ancienne, nouvelle in champs:
            if str(ancienne) != str(nouvelle):
                cur.execute("""
                    INSERT INTO historique_modifications (article_id, champ_modifie, ancienne_valeur, nouvelle_valeur)
                    VALUES (%s, %s, %s, %s)
                """, (id, champ, ancienne, nouvelle))

        cur.execute("""
            UPDATE articles SET nom=%s, description=%s, prix=%s, stock=%s, categorie_id=%s WHERE id=%s
        """, (nom, description, prix, stock, categorie_id, id))
        conn.commit()
        cur.close()
        conn.close()
        flash('Article modifié avec succès !', 'success')
        return redirect(url_for('index'))

    cur.execute("SELECT * FROM articles WHERE id = %s", (id,))
    article = cur.fetchone()
    cur.execute("SELECT * FROM categories")
    categories = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('modifier.html', article=article, categories=categories)

@app.route('/historique/<int:id>')
@login_required
def historique_article(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM articles WHERE id = %s", (id,))
    article = cur.fetchone()
    cur.execute("""
        SELECT * FROM historique_modifications 
        WHERE article_id = %s 
        ORDER BY date_modification DESC
    """, (id,))
    historique = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('historique.html', article=article, historique=historique)

# SUPPRIMER UN ARTICLE
@app.route('/supprimer/<int:id>')
@login_required
def supprimer(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM historique_modifications WHERE article_id = %s", (id,))
    cur.execute("DELETE FROM ventes WHERE article_id = %s", (id,))
    cur.execute("DELETE FROM articles WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Article supprimé !', 'danger')
    return redirect(url_for('index'))

# PAGE PUBLIQUE - Scan QR code (pas de login_required ici !)
@app.route('/article/<int:id>')
def article_public(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM articles WHERE id = %s", (id,))
    article = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('article_public.html', article=article)

# ENREGISTRER UNE VENTE (pas de login_required ici !)
@app.route('/vendre/<int:id>', methods=['POST'])
def vendre(id):
    quantite = int(request.form['quantite'])
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT stock FROM articles WHERE id = %s", (id,))
    article = cur.fetchone()

    if article and article[0] >= quantite:
        cur.execute("UPDATE articles SET stock = stock - %s WHERE id = %s", (quantite, id))
        cur.execute("INSERT INTO ventes (article_id, quantite) VALUES (%s, %s)", (id, quantite))
        conn.commit()
        flash('success', 'vente_ok')
    else:
        flash('Stock insuffisant !', 'danger')

    cur.close()
    conn.close()
    return redirect(url_for('article_public', id=id))

@app.route('/export/stock')
@login_required
def export_stock():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.nom, c.nom as categorie, a.prix, a.stock, (a.prix * a.stock) as valeur_stock, a.date_ajout
        FROM articles a
        LEFT JOIN categories c ON a.categorie_id = c.id
        ORDER BY a.nom
    """)
    articles = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    # Style en-tête
    header_fill = PatternFill(start_color="2C1A0E", end_color="2C1A0E", fill_type="solid")
    header_font = Font(color="D4A843", bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['Article', 'Catégorie', 'Prix (FCFA)', 'Stock', 'Valeur Stock (FCFA)', 'Date Ajout']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Données
    for row, article in enumerate(articles, 2):
        for col, value in enumerate(article, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = center
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="FDF6E8", end_color="FDF6E8", fill_type="solid")

    # Largeur colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True, download_name='stock_saahel.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/ventes')
@login_required
def export_ventes():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.nom, c.nom as categorie, v.quantite, a.prix, (v.quantite * a.prix) as total, v.date_vente
        FROM ventes v
        JOIN articles a ON v.article_id = a.id
        LEFT JOIN categories c ON a.categorie_id = c.id
        ORDER BY v.date_vente DESC
    """)
    ventes = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"

    header_fill = PatternFill(start_color="2C1A0E", end_color="2C1A0E", fill_type="solid")
    header_font = Font(color="D4A843", bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['Article', 'Catégorie', 'Quantité', 'Prix Unitaire (FCFA)', 'Total (FCFA)', 'Date Vente']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row, vente in enumerate(ventes, 2):
        for col, value in enumerate(vente, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = center
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="FDF6E8", end_color="FDF6E8", fill_type="solid")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True, download_name='ventes_saahel.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# TABLEAU DE BORD
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM articles")
    total_articles = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM ventes")
    total_ventes = cur.fetchone()[0]

    cur.execute("SELECT SUM(quantite) FROM ventes")
    total_quantite = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(v.quantite * a.prix) FROM ventes v JOIN articles a ON v.article_id = a.id")
    chiffre_affaires = cur.fetchone()[0] or 0

    cur.execute("SELECT * FROM articles WHERE stock <= 5")
    stock_bas = cur.fetchall()

    cur.execute("""
        SELECT a.nom, SUM(v.quantite) as total
        FROM ventes v JOIN articles a ON v.article_id = a.id
        GROUP BY a.id ORDER BY total DESC LIMIT 5
    """)
    top_articles = cur.fetchall()

    cur.execute("""
        SELECT v.id, a.nom, v.quantite, a.prix, (v.quantite * a.prix), v.date_vente
        FROM ventes v JOIN articles a ON v.article_id = a.id
        ORDER BY v.date_vente DESC LIMIT 50
    """)
    historique = cur.fetchall()

    # Courbe ventes par jour (30 derniers jours)
    cur.execute("""
        SELECT DATE(date_vente) as jour, SUM(quantite) as total
        FROM ventes
        WHERE date_vente >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        GROUP BY DATE(date_vente)
        ORDER BY jour
    """)
    ventes_jour_raw = cur.fetchall()

    # Chiffre d'affaires par semaine (8 dernières semaines)
    cur.execute("""
        SELECT YEARWEEK(date_vente, 1) as semaine,
               MIN(DATE(date_vente)) as debut,
               SUM(v.quantite * a.prix) as ca
        FROM ventes v JOIN articles a ON v.article_id = a.id
        WHERE date_vente >= DATE_SUB(NOW(), INTERVAL 56 DAY)
        GROUP BY YEARWEEK(date_vente, 1)
        ORDER BY semaine
    """)
    ca_semaine_raw = cur.fetchall()

    # Ventes par article (top 8)
    cur.execute("""
        SELECT a.nom, SUM(v.quantite) as total
        FROM ventes v JOIN articles a ON v.article_id = a.id
        GROUP BY a.id ORDER BY total DESC LIMIT 8
    """)
    ventes_article_raw = cur.fetchall()

    cur.close()
    conn.close()

    # Formater pour Chart.js
    ventes_jour_labels = [str(r[0]) for r in ventes_jour_raw]
    ventes_jour_data = [int(r[1]) for r in ventes_jour_raw]

    ca_semaine_labels = [f"Sem. {str(r[1])}" for r in ca_semaine_raw]
    ca_semaine_data = [float(r[2]) for r in ca_semaine_raw]

    ventes_article_labels = [r[0] for r in ventes_article_raw]
    ventes_article_data = [int(r[1]) for r in ventes_article_raw]

    return render_template('dashboard.html',
        total_articles=total_articles,
        total_ventes=total_ventes,
        total_quantite=total_quantite,
        chiffre_affaires=chiffre_affaires,
        stock_bas=stock_bas,
        top_articles=top_articles,
        historique=historique,
        ventes_jour_labels=ventes_jour_labels,
        ventes_jour_data=ventes_jour_data,
        ca_semaine_labels=ca_semaine_labels,
        ca_semaine_data=ca_semaine_data,
        ventes_article_labels=ventes_article_labels,
        ventes_article_data=ventes_article_data
    )

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))